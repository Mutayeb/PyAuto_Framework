# Read the CSV or EXCEL file
# Create a Function create_token which can take values from the Excel File
# Verify the Expected Result.

# Read the Excel - openpyxl
import openpyxl

from src.constants.apiconstants import apiConstants
from src.utils.utils import Util
from src.helpers.api_requests_wrapper import *
from src.helpers.common_verification import *

def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(filename=file_path) #The load_workbook() function opens file for reading
    sheet1 = workbook.active #The object of the workbook.active has been created in the script to read the values of the max_row and the max_column properties.
    #for loop to read each row from the excel sheet
    for row in sheet1.iter_rows(min_row=2,values_only=True):
        username,password = row
        credentials.append({"username":username,
                            "password":password})
    return credentials

def create_auth_request(username,password):
        payload = {"username":username,
                   "password":password}
        response = post_request(
            url=apiConstants.url_create_token(),
            headers=Util().common_headers_json(),
            auth=None,
            payload=payload,
            in_json=False
        )
        return response

def test_create_auth_with_excel():
        file_path ="C:/Users/mutay/OneDrive/Documents/courses/PyATB/code files/pyProject1-AutomationFramework/tests/test/datadriventesting/testdata_ddt.xlsx"
        credentials = read_credentials_from_excel(file_path=file_path)
        print(credentials)
        #Using for loop to send requests with all the data(each row) from the excel sheet
        for user_cred in credentials:
            username = user_cred["username"]
            password = user_cred["password"]
            print(username, password)
            response = create_auth_request(username=username, password=password)
            print(response.status_code)
            verify_status_code(response_data=response,expect_data=200)